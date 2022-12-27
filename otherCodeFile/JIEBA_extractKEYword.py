from tqdm import tqdm
import jieba.analyse
import pandas
def readXls(path=r'C:\Users\glodon\Desktop\jobCollect\zhilian\智联造价员0.xlsx'):
    # 打开文件：
    from openpyxl import load_workbook
    excel = load_workbook(path)
    # 获取sheet：
    table = excel.get_sheet_by_name('Sheet2')  # 通过表名获取
    # 获取行数和列数：
    rows = table.max_row  # 获取行数
    cols = table.max_column  # 获取列数
    jobNames=[]
    jobDess=[]
    for i in range(2,rows+1):
        jobName=table.cell(row=i, column=1).value
        jobDes=table.cell(row=i, column=2).value
        if jobDes not in jobDess:
            jobNames.append(jobName)
            jobDess.append(jobDes)
    return jobNames,jobDess
def tfidfDeal(jobs=[]):
    from sklearn.feature_extraction.text import CountVectorizer, TfidfTransformer,TfidfVectorizer
    text = jobs[:1]
    singeText=text

    # text=["I come to China to travel",
    #     "This is a car polupar in China",
    #     "I love tea and Apple ",
    #     "The work is to write some papers in science"]

    # text=["岗位职责:1、大专及以上学历，工程造价专业1年及以上工作经验；2、熟练应用广联达等造价软件，以及office、autocad等相关的办公软件；"+
    #       "能独立完成:编审工程项目概算、预算、结算、竣工决算、工程审核、工程量清单、标底及对量等工作；3、工作严谨，善于沟通，具备良好的团队合作精神和职业操守；"+
    #       "4、卓越的执行能力，学习能力和独立工作能力。5、掌握新技术，了解新材料和国内工程造价动态；6、责任心强，有敬业精神，能吃苦耐劳，品质优良。任职要求："+
    #       "1.五官端正，非民办正规院校毕业相关专业大专及以上学历，招投标、造价等工作，上岗证一项以上，符合条件者可来应聘；2.具有较强的工作责任感和事业心；"
    #       "3.具有良好的专业素质和职业化素养。4.具有注册类证书者优先。"
    # ,
    #     "协助造价工程师完成工程项目。岗位要求：专业：土建、工程管理、安装、水暖、工程造价类相关专业；学历：大学专科，本科，应届毕业生；其他："
    #     "a. 具备较强的责任心、语言表达、沟通协作能力。b.积极主动、有上进心、求知欲；c.有社会实践经验。d.抗压能力强"
    # ]

    tfidf=TfidfVectorizer()
    weight=tfidf.fit_transform(singeText).toarray()
    word=tfidf.get_feature_names()
    print('单词数量', len(word),word)

    # print(weight)

    # for i in range(len(weight)):
    #     print('第{}句，关键词权重'.format(i))
    #     for j in range(len(word)):
    #         print(word[j],weight[i][j])

def jiebaDeal(jobs=[]):
    txt=jobs[0][1]
    keywords=jieba.analyse.extract_tags(txt, topK=100, withWeight=False, allowPOS=())
    print(len(keywords),keywords)
if __name__ == '__main__':
    _,text=readXls()
    # jiebaDeal(jobs=text)
    tfidfDeal(jobs=text)



